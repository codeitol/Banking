import os
from os.path import exists
import re
import openpyxl
if not exists("Data.xlsx"):
      wb=openpyxl.Workbook()
      wb.create_sheet('Cdetails',0)
      wb.create_sheet('Account',1)
      wb.save("Data.xlsx")
wb = openpyxl.load_workbook('Data.xlsx')
ws1 = wb['Cdetails']
ws2=wb['Account']
z=['A','B','C','D']
for j in z:
    ws1.column_dimensions[j].width=20
    ws2.column_dimensions[j].width=20
ws1["A1"]="First Name"
ws1["B1"]="Last Name"
ws1["C1"]="User Name"
ws1["D1"]="Password"
ws2["A1"]="User Name"
ws2["B1"]="Debit"
ws2["C1"]="Credit"
ws2["D1"]="Balance"
w=os.get_terminal_size().columns
x=5
wb.save("Data.xlsx")
print("WELCOME!! TO NET BANKING".center(w))
def signup():
     print("WELCOME!! Kindly Fill The Details".center(w))
     fn=input("Enter First Name:")
     ln=input("Enter Last Name:")
     un=input("Enter User Name:")
     fg=0
     while fg==0:
          ps=input("Generate Your password:")
          if((len(ps))>7)and(re.search("[A-Z]",ps))and(re.search("[a-z]",ps))and(re.search("[_@&$]",ps))and(re.search("[0-9]",ps)):
               fg=1
          else:
              print("Error!! Please Fill Again")
     em=0
     while em==0:
           el=input("Enter E-mail:")
           if((re.search("[@]",el))and(re.search("[.]",el))):
                em=1
           else:
                print("Error!! Please Fill Again")
     r = ws1.max_row
     c = ws1.max_column
     k=[fn,ln,un,ps]
     s=[un,0,0,10000]
     for i in range(r+1, r+2):
          for j in range(1, c+1):
     	       ws1.cell(row=i, column=j).value=k[j-1]
     e=ws2.max_row
     d=ws2.max_column
     for h in range(e+1,e+2):
          for f in range(1,d+1):
               ws2.cell(row=h, column=f).value=s[f-1]
     wb.save("Data.xlsx")
def help(uns):
      r=ws2.max_row 
      for i in range(r,1,-1):
           if(uns==ws2.cell(row=i,column=1).value):
                  return(i)
def credit(uns,r):
     q=int(input("Enter Amount To Be Credited:"))
     i=help(uns)
     ws2.cell(row=r+1,column=4).value=int(ws2.cell(row=i,column=4).value)+q
     ws2.cell(row=r+1,column=1).value=uns
     ws2.cell(row=r+1,column=3).value=q
def debit(uns,r):
     q=int(input("Enter Amount To Be Debited:"))
     i=help(uns)
     ws2.cell(row=r+1,column=4).value=int(ws2.cell(row=i,column=4).value)-q
     ws2.cell(row=r+1,column=1).value=uns
     ws2.cell(row=r+1,column=2).value=q
def chng(uns,r):
     fg=0
     while fg==0:
          ps=input("Enter New Password:")
          if((len(ps))>7)and(re.search("[A-Z]",ps))and(re.search("[a-z]",ps))and(re.search("[_@&$]",ps))and(re.search("[0-9]",ps)):
               fg=1
          else:
              print("Error!! Please Fill Again")
     for i in range(2,r+1):
           if(uns==ws1.cell(row=i,column=3).value):
                 ws1.cell(row=i,column=4).value=ps
def tran(uns,r):
     w=input("Enter Username:")
     q=int(input("Enter Amount To Be Transferred:"))
     i=help(uns)
     ws2.cell(row=r+1,column=4).value=int(ws2.cell(row=i,column=4).value)-q
     ws2.cell(row=r+1,column=1).value=uns
     ws2.cell(row=r+1,column=2).value=q
     i=help(w)  
     ws2.cell(row=r+2,column=4).value=int(ws2.cell(row=i,column=4).value)+q
     ws2.cell(row=r+2,column=1).value=w
     ws2.cell(row=r+2,column=3).value=q        
def signin():
     print("WELCOME!! Please Sign-In".center(w))
     uns=input("Enter User Name:")
     r=ws1.max_row
     v=ws2.max_row
     for i in range(2,r+1):
           if(uns==ws1.cell(row=i,column=3).value):
                 pss=input("Enter Your password:")
                 if(pss==ws1.cell(row=i,column=4).value):
                      print("WELCOME!!".center(w))
                      p=6
                      while p!=5:
                            p=int(input("Please choose from the following:\n1.Credit\n2.Debit\n3.Change password\n4.Transfer\n5.Log Out"))
                            if(p==1):
                               credit(uns,v)
                            elif(p==2):
                               debit(uns,v)
                            elif(p==3):
                               chng(uns,r)
                            elif(p==4):
                               tran(uns,v) 
                            else:
                               print("Thank you")
                            wb.save("Data.xlsx")
                 else:
                     print("Password Incorrect !! Please Try Again")                
while x!=0:
    x=int(input("Please choose from the following:\n1.Sign Up\n2.Sign In\n0.Exit\n"))
    if(x==1):
        signup()
    elif(x==2):
        signin()
    elif(x!=0):
        print("Invalid input TRY AGAIN!!")
    else:
        print("Thank You!!")
